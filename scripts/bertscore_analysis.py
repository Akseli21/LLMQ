# ============================================================
# bertscore_analysis.py
# Analyse BERTScore — Qualité, Stabilité, Reformulation
#
# Auteur : Julia CANARELLI — ESME Paris ING2 2025-2026
# Projet : LLMQ — Évaluation des performances d'un LLM
#          en contexte d'aide au diagnostic clinique
#
# Usage : python scripts/bertscore_analysis.py
# Input : GrilleFinale_Evaluation_LLMQ_rempli_V2.xlsx
# Output: results/bertscore_resultats.xlsx
# ============================================================

import sys
from collections import defaultdict
from itertools import combinations
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration — adapter les chemins selon votre machine
# ---------------------------------------------------------------------------
BASE_DIR    = Path(".")
INPUT_FILE  = BASE_DIR / "results" / "GrilleFinale_Evaluation_LLMQ_rempli_V2.xlsx"
OUTPUT_FILE = BASE_DIR / "results" / "bertscore_resultats.xlsx"

SHEET_MODELS = {
    "EXPERT   LLaMa 3.1 8B Instruct ": "LLaMa 3.1 8B Instruct",
    "EXPERT Gemma3 27b":                "Gemma3 27b",
    "EXPERT qwen2.5_32b":               "qwen2.5:32b",
}

COL_CASE_ID   = 0
COL_PROMPT_ID = 1
COL_REP_NUM   = 3
COL_REPONSE   = 6
COL_REFERENCE = 7

BERT_LANG = "fr"

THRESH = {
    "brut":    {"green": 0.85, "yellow": 0.70},
    "rescale": {"green": 0.50, "yellow": 0.20},
}

FILL_GREEN    = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW   = PatternFill("solid", fgColor="FFEB9C")
FILL_RED      = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER   = PatternFill("solid", fgColor="305496")
FILL_SUBHEADER= PatternFill("solid", fgColor="8EA9DB")
FILL_BRUT     = PatternFill("solid", fgColor="D9E1F2")
FILL_RESCALE  = PatternFill("solid", fgColor="FCE4D6")
FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_SUBHEADER= Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def fill_for_score(score, mode):
    if score is None:
        return None
    t = THRESH[mode]
    if score >= t["green"]:
        return FILL_GREEN
    if score >= t["yellow"]:
        return FILL_YELLOW
    return FILL_RED


def load_rows(workbook):
    out = {}
    for sheet_name in SHEET_MODELS:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Feuille manquante : '{sheet_name}'")
        ws = workbook[sheet_name]
        rows = []
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r[COL_CASE_ID] is None:
                continue
            rows.append({
                "case_id":   str(r[COL_CASE_ID]).strip(),
                "prompt_id": str(r[COL_PROMPT_ID]).strip(),
                "rep_num":   int(r[COL_REP_NUM]),
                "reponse":   (r[COL_REPONSE]   or "").strip(),
                "reference": (r[COL_REFERENCE] or "").strip(),
            })
        out[sheet_name] = rows
        print(f"  '{sheet_name}': {len(rows)} lignes lues")
    return out


_SCORERS = {}


def get_scorer(mode):
    global _SCORERS
    if mode not in _SCORERS:
        from bert_score import BERTScorer
        _SCORERS[mode] = BERTScorer(lang=BERT_LANG, rescale_with_baseline=(mode == "rescale"))
    return _SCORERS[mode]


def compute_bertscore(cands, refs, mode):
    if not cands:
        return [], [], []
    cands = [c if c.strip() else " " for c in cands]
    refs  = [r if r.strip() else " " for r in refs]
    P, R, F1 = get_scorer(mode).score(cands, refs)
    return P.tolist(), R.tolist(), F1.tolist()


def analyse_qualite(rows):
    cands = [r["reponse"]   for r in rows]
    refs  = [r["reference"] for r in rows]
    out = {}
    for mode in ("brut", "rescale"):
        P, R, F1 = compute_bertscore(cands, refs, mode)
        out[mode] = [
            {"case_id": row["case_id"], "prompt_id": row["prompt_id"],
             "rep_num": row["rep_num"], "reference": row["reference"],
             "precision": p, "rappel": r_, "f1": f}
            for row, p, r_, f in zip(rows, P, R, F1)
        ]
    return out


def _pair_analysis(rows, group_key, pair_label):
    by_group = defaultdict(list)
    for r in rows:
        by_group[group_key(r)].append(r)
    cands, refs, meta = [], [], []
    for grp_key, group in by_group.items():
        group = sorted(group, key=pair_label["sort"])
        for a, b in combinations(group, 2):
            cands.append(a["reponse"])
            refs.append(b["reponse"])
            meta.append(pair_label["meta"](a, b, grp_key))
    out = {}
    for mode in ("brut", "rescale"):
        P, R, F1 = compute_bertscore(cands, refs, mode)
        out[mode] = [{**m, "precision": p, "rappel": r_, "f1": f}
                     for m, p, r_, f in zip(meta, P, R, F1)]
    return out


def analyse_stabilite(rows):
    return _pair_analysis(
        rows,
        group_key=lambda r: (r["case_id"], r["prompt_id"]),
        pair_label={
            "sort": lambda x: x["rep_num"],
            "meta": lambda a, b, g: {
                "case_id": g[0], "prompt_id": g[1],
                "paire": f"rep{a['rep_num']} vs rep{b['rep_num']}",
            },
        },
    )


def analyse_reformulation(rows):
    return _pair_analysis(
        rows,
        group_key=lambda r: (r["case_id"], r["rep_num"]),
        pair_label={
            "sort": lambda x: x["prompt_id"],
            "meta": lambda a, b, g: {
                "case_id": g[0], "rep_num": g[1],
                "paire": f"{a['prompt_id']} vs {b['prompt_id']}",
            },
        },
    )


def main():
    print("=" * 70)
    print("BERTScore LLMQ — Endocrinologie (BRUT + RESCALÉ)")
    print("=" * 70)

    if not INPUT_FILE.exists():
        print(f"ERREUR : fichier introuvable {INPUT_FILE}")
        sys.exit(1)

    print("\n[1/3] Lecture du fichier source")
    wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    data  = load_rows(wb_in)

    synth = {
        "brut":    {"qualite": {}, "stabilite": {}, "reformulation": {}},
        "rescale": {"qualite": {}, "stabilite": {}, "reformulation": {}},
    }

    print("\n[2/3] Calcul des BERTScores")
    for sheet_name, model_label in SHEET_MODELS.items():
        print(f"\n  Modèle : {model_label}")
        rows  = data[sheet_name]
        qual  = analyse_qualite(rows)
        stab  = analyse_stabilite(rows)
        reform= analyse_reformulation(rows)

        for mode in ("brut", "rescale"):
            sums_q = defaultdict(list)
            for q in qual[mode]:
                sums_q[q["case_id"]].append(q["f1"])
            synth[mode]["qualite"][model_label] = {k: sum(v)/len(v) for k,v in sums_q.items()}

            sums_s = defaultdict(list)
            for s in stab[mode]:
                sums_s[s["case_id"]].append(s["f1"])
            synth[mode]["stabilite"][model_label] = {k: sum(v)/len(v) for k,v in sums_s.items()}

            sums_r = defaultdict(list)
            for r in reform[mode]:
                sums_r[r["case_id"]].append(r["f1"])
            synth[mode]["reformulation"][model_label] = {k: sum(v)/len(v) for k,v in sums_r.items()}

    print("\n\n" + "=" * 90)
    print("RÉSULTATS BERTSCORE")
    print("=" * 90)

    for model_label in SHEET_MODELS.values():
        print(f"\nMODÈLE : {model_label}")
        print("-" * 60)
        for mode in ("brut", "rescale"):
            print(f"  Mode : {mode.upper()}")
            for dim in ("qualite", "stabilite", "reformulation"):
                vals = list(synth[mode][dim][model_label].values())
                mean = sum(vals) / len(vals) if vals else 0
                print(f"    F1 moyen {dim:15s} : {mean:.3f}")

    print("\n[3/3] Export Excel →", OUTPUT_FILE)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.save(OUTPUT_FILE)
    print("Terminé.")


if __name__ == "__main__":
    main()
